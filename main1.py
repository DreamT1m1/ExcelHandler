import openpyxl
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.common import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support import expected_conditions as EC

from selenium.webdriver.support.wait import WebDriverWait

import os


WORKBOOK_NAME = "madis.xlsx"
USED_COLUMNS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

# Load an Excel workbook
work_book = openpyxl.load_workbook(WORKBOOK_NAME)
ws = work_book.active

# Set a searchable color
SEARCHABLE_COLOR = "FFFFFF00"  # Yellow

#Get an Excel table rows as list
rows_list = list(ws.iter_rows())

# Get Excel table length
table_len = len(rows_list)

# Get amount of rows with searchable color
searchable_color_rows_amount = sum(1 for i in range(1, table_len + 1)
                                   if ws[f"A{i}"].fill.start_color.rgb == SEARCHABLE_COLOR)

print(f"Got {searchable_color_rows_amount} yellow rows\n")

# Set the web driver
driver = webdriver.Chrome()

# How many devices was processed successfully
successfully_processed = 0

# Iterate the rows
for i in range(1, table_len + 1):
    # Check if the row is SEARCHABLE color
    if ws[f"A{i}"].fill.start_color.rgb == SEARCHABLE_COLOR:
        # Get RP-code which is going to be used for search in MADIS
        searched_rp_code = str(ws[f"A{i}"].value)

        # Enter the main MADIS page
        driver.get("http://tehnika/hooldus/toolaud")

        try:
            # Get the login button
            button = WebDriverWait(driver, 1).until(
                EC.element_to_be_clickable((By.ID, "sisselogimisenupp"))
            )
            button.click()
            print("Entered Successfully\n")
        except TimeoutException:
            pass

        print(f"Processing a device with RP-code: {searched_rp_code}. Column: A{i}")

        # Get the search field
        search_field = driver.find_element(By.CSS_SELECTOR, "input.form-control")
        # Clear the field
        search_field.clear()
        # Past the searched RP-code
        search_field.send_keys(searched_rp_code.strip())
        # "Press ENTER button"
        search_field.send_keys(Keys.RETURN)
        time.sleep(2)

        # Get a table of found devices
        try:
            table = driver.find_element(By.ID, "sugseade-tab")
        except NoSuchElementException:
            try:
                table = driver.find_element(By.ID, "sugseade-div")
            except NoSuchElementException:
                continue
            else:
                print("Table of devices found!")
        else:
            print("Table of devices found!")

        # Get rows of the table
        rows = table.find_elements(By.TAG_NAME, "tr")
        if len(rows) == 0:
            continue
        # Iterate the rows
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            if searched_rp_code in [i.text for i in cells]:
                row.click()
                print("Device found!")
                break

        # Check if it is a complex of devices
        try:
            multiple_devices = driver.find_element(By.ID, "komplekt")
        except NoSuchElementException:
            is_complex = False
        else:
            is_complex = True

        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        info = driver.find_elements(By.TAG_NAME, "b")

        # Search for the main info html element
        main_info_element = driver.find_element(By.ID, "seadmeylevaade")
        # Search for the name of the device
        name_in_system_h2 = main_info_element.find_element(By.TAG_NAME, "h2").text
        print("Name in 'Madis':", name_in_system_h2)

        for el in info:
            parent = el.find_element(By.XPATH, "..")
            full_text = parent.text.strip()
            data = dict(filter(lambda x: len(x) == 2, [tuple(i.split(": ")) for i in full_text.split("\n")]))
            device_ltkh_code = ''
            device_sn = ''
            supervisor = ''
            place = ''

            complex_sn = ''
            complex_code = ''
            complex_rp_code = ''

            # Check if it is a complex
            if not is_complex:
                # Search for the either LTKH code or serial number
                if data["RP-kood"] == searched_rp_code:
                    if "Seadme nr" in data:
                        device_ltkh_code = data["Seadme nr"]
                        print("LKTH code", device_ltkh_code)
                    else:
                        print("No device LTKH code found")
                        device_sn = data["Seerianr"]
                        print("Seerianr", device_sn)
            else:
                # TODO
                print("Multiple devices")
                complex_sn = data.get("Seerianr", "")
                complex_code = data.get("Seadme nr", "")
                complex_rp_code = data.get("RP-kood", "")
                if complex_sn:
                    print(f"S/n of the complex: {complex_sn}")
                if complex_code:
                    print(f"LTKH code of the complex: {complex_code}")
                if complex_rp_code:
                    print(f"RP-code of the complex: {complex_rp_code}")

                driver.get(" ") # TODO: Get a page with exact part of the complex and get all the data of the part
                additional_search_field = driver.find_element(
                    By.ID, " ") # TODO: Get a HTML element of search bar and then insert
                # Clear the field
                additional_search_field.clear()
                # Inset the value
                additional_search_field.send_keys(searched_rp_code)
                # 'Press' the Enter button
                additional_search_field.send_keys(Keys.RETURN)
                WebDriverWait(driver, 2)
                # TODO: process the result. Get the row of found table with searched complex part

            # Search for the device's supervisor
            if "Vastutaja" in data:
                supervisor = data["Vastutaja"].replace(" (Muuda)", "").replace("(Muuda)", "")
                print(f"Supervisor: {supervisor}")
            # Search for the place where the device is
            if "Asukoht" in data:
                place = data["Asukoht"].replace(" (Muuda)", "")
                if place.strip() != '/':
                    print(f"Placed in: {place}")

            # Enter the name as it is in the 'Madis'
            ws[f"C{i}"].value = name_in_system_h2

            if not is_complex:
                # Enter supervisor's name if it is in the 'Madis'
                if supervisor:
                    ws[f"E{i}"].value = supervisor
                # Enter LTKH code if it is in the 'Madis'
                if device_ltkh_code:
                    ws[f"F{i}"].value = device_ltkh_code
                # Enter device serial number in case it was found
                if device_sn:
                    ws[f"G{i}"].value = f"s/n {device_sn}"
                # Enter the place where the device is
                if place.strip() != "/":
                    ws[f"H{i}"].value = place
            else:
                # Enter the complex s/n
                if complex_sn:
                    ws[f"G{i}"].value = f"Komplekti s/n: {complex_sn}"
                if complex_code:
                    ws[f"G{i}"].value = f"Komplekti LTKH kood: {complex_code}"

                # TODO: Enter all the data of the exact complex part

            # Set a new fill color
            fill = PatternFill(start_color="FF04D3FC",
                               end_color="FF04D3FC",
                               fill_type="solid")

            # Change the color of processed row
            for col in USED_COLUMNS:
                ws[f"{col}{i}"].fill = fill
            # Save the workbook
            try:
                work_book.save("madis.xlsx")
            except PermissionError:
                print("Looks like the file is opened right now.\nWaiting until the file is closed...")
                lock_file = "~$" + WORKBOOK_NAME

                while os.path.exists(lock_file):
                    time.sleep(1)
                work_book.save("madis.xlsx")
            break

        successfully_processed += 1

        print(f"Currently processed "
              f"{successfully_processed} of {searchable_color_rows_amount}: "
              f"{round(successfully_processed / searchable_color_rows_amount * 100, 3)}%\n")

work_book.close()
driver.quit()
